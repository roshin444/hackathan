import streamlit as st
import pandas as pd
from io import BytesIO
import json
from risk_agent import RiskAssessmentAgent
from config import llm
from langchain.schema import SystemMessage, HumanMessage
import plotly.express as px
from openpyxl.styles import PatternFill, Font
import time

st.set_page_config(page_title="Audit & Compliance Risk Agent", layout="wide")
st.title("Audit & Compliance Risk Assessment Portal")

# --- Project Banner & About ---
st.image("banner.jpg", use_column_width=True)
st.markdown("""
### About This Project
This portal allows organizations to perform automated **Audit & Compliance Risk Assessment**.
Upload your audit findings, compliance checklists, and risk criteria to generate:
- Risk summaries with priority-based classification
- Interactive Q&A with AI expert insights
- Graphical dashboards for visual analysis
- Performance metrics of the risk assessment process
""")

tabs = st.tabs(["Risk Assessment Summary", "Interactive Q&A", "Risk Dashboard", "Performance Metrics"])
summary_tab, qna_tab, dashboard_tab, perf_tab = tabs

# ---------------- Risk Assessment Summary ----------------
with summary_tab:
    st.subheader("Upload Required Files")
    audit_file = st.file_uploader("Upload audit findings CSV", type=["csv"])
    checklist_file = st.file_uploader("Upload compliance checklist CSV", type=["csv"])
    criteria_file = st.file_uploader("Upload risk criteria JSON", type=["json"])

    if audit_file and checklist_file and criteria_file:
        try:
            audit_df = pd.read_csv(audit_file)
            checklist_df = pd.read_csv(checklist_file)
            criteria_json = json.load(criteria_file)
            agent = RiskAssessmentAgent(
                findings_path_or_df=audit_df,
                checklist_path_or_df=checklist_df,
                criteria_path_or_dict=criteria_json
            )
        except Exception as e:
            st.error(f"Error initializing agent: {e}")
            st.stop()

        if st.button("Run Risk Assessment"):
            with st.spinner("Generating risk assessment..."):
                start_time = time.time()
                try:
                    summary_list = agent.generate_summary()
                    if not summary_list:
                        st.warning("No findings returned. Check inputs.")
                    else:
                        st.session_state["report_df"] = pd.DataFrame(summary_list)
                        st.session_state["summary_time_secs"] = time.time() - start_time
                except Exception as e:
                    st.error(f"Error generating summary: {e}")
                    st.stop()

        if "report_df" in st.session_state and not st.session_state["report_df"].empty:
            df = st.session_state["report_df"].copy()

            # Ensure core columns exist with safe defaults
            for c, d in [("risk_level", "Low"), ("status", "Open"),
                         ("severity", "Medium"), ("finding_id", "N/A")]:
                if c not in df.columns:
                    df[c] = d

            # Dynamic description mapping
            desc_candidates = ["description", "Finding", "control_text", "test_procedure", "Area", "area", "finding"]
            for cand in desc_candidates:
                if cand in df.columns:
                    df["description"] = df[cand]
                    break
            if "description" not in df.columns:
                df["description"] = "N/A"

            # Filters
            st.markdown("### 🔍 Filters")
            col1, col2 = st.columns(2)
            with col1:
                risk_choices = sorted(df["risk_level"].dropna().unique().tolist())
                risk_filter = st.multiselect("Filter by Risk Level", options=risk_choices, default=risk_choices)
            with col2:
                status_choices = sorted(df["status"].dropna().unique().tolist()) if "status" in df.columns else []
                status_filter = st.multiselect("Filter by Status", options=status_choices, default=status_choices)

            filtered_df = df[df["risk_level"].isin(risk_filter)]
            if status_filter:
                filtered_df = filtered_df[filtered_df["status"].isin(status_filter)]

            # Priority sort
            priority_map = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
            filtered_df["priority_order"] = filtered_df["risk_level"].map(priority_map).fillna(0)
            filtered_df = filtered_df.sort_values(by="priority_order", ascending=False)

            # Styling function
            def highlight_risk(row):
                cmap = {
                    "Critical": "background-color: red; color: white;",
                    "High": "background-color: orange; color: black;",
                    "Medium": "background-color: yellow; color: black;",
                    "Low": "background-color: green; color: white;"
                }
                return [cmap.get(row["risk_level"], "") if col == "risk_level" else "" for col in row.index]

            st.subheader("Merged Findings Summary")
            if "Area" in filtered_df.columns:
                for area in filtered_df["Area"].unique():
                    with st.expander(f"{area} Findings"):
                        area_df = filtered_df[filtered_df["Area"] == area]
                        st.dataframe(area_df.style.apply(highlight_risk, axis=1), use_container_width=True)
            else:
                st.dataframe(filtered_df.style.apply(highlight_risk, axis=1), use_container_width=True)

            # Alerts
            if "status" in filtered_df.columns:
                critical_open = filtered_df[(filtered_df["risk_level"].isin(["Critical", "High"])) &
                                            (filtered_df["status"].str.lower() == "open")]
                if not critical_open.empty:
                    st.warning(f"⚠️ {len(critical_open)} open High/Critical risks!")
                    show_cols = [c for c in ["finding_id", "description", "risk_level", "status"] if c in filtered_df.columns]
                    st.table(critical_open[show_cols])
                else:
                    st.success("No immediate High/Critical open risks.")
            else:
                st.info("No status column detected; skipping live alerts.")

            # Excel export of the filtered view
            excel_buf = BytesIO()
            with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                filtered_df.to_excel(writer, index=False, sheet_name="Audit_Report")
                ws = writer.sheets["Audit_Report"]
                color_map = {"Critical": "FF0000", "High": "FFA500", "Medium": "FFFF00", "Low": "00FF00"}
                for row_idx, risk in enumerate(filtered_df["risk_level"], start=2):
                    try:
                        col_idx = filtered_df.columns.get_loc("risk_level") + 1
                        cell = ws.cell(row=row_idx, column=col_idx)
                        fill_col = color_map.get(risk, "FFFFFF")
                        cell.fill = PatternFill(start_color=fill_col, end_color=fill_col, fill_type="solid")
                        if risk in ["Critical", "High"]:
                            cell.font = Font(bold=True, color="000000")
                    except Exception:
                        pass
            excel_buf.seek(0)
            st.download_button("Download Filtered Excel Report", excel_buf,
                               "audit_compliance_filtered_report.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- Interactive Q&A ----------------
with qna_tab:
    st.subheader("Interactive Chat on Merged Report")
    if "chat_history" not in st.session_state:
        st.session_state["chat_history"] = []

    user_question = st.chat_input("Ask a question about the report...")
    if user_question:
        st.session_state["chat_history"].append({"role": "user", "message": user_question})
        context = (st.session_state["report_df"].to_json(orient="records")
                   if "report_df" in st.session_state else "No report data available.")
        messages = [
            SystemMessage(content="You are an audit & compliance risk assessment expert. Answer concisely based on the report."),
            HumanMessage(content=f"Report Data: {context}\n\nQuestion: {user_question}")
        ]
        with st.spinner("Generating answer..."):
            try:
                resp = llm.invoke(messages)
                ai_text = resp.content
            except Exception as e:
                ai_text = f"Error generating answer: {e}"
        st.session_state["chat_history"].append({"role": "ai", "message": ai_text})

    # Display chat history with thumbs up/down
    for idx, msg in enumerate(st.session_state["chat_history"]):
        if msg["role"] == "user":
            st.markdown(f"**User:** {msg['message']}")
        else:
            left, right = st.columns([0.9, 0.1])
            with left:
                st.markdown(f"**AI:** {msg['message']}")
            with right:
                col_key = f"feedback_{idx}"
                if st.button("👍", key=f"up_{idx}"):
                    st.session_state[f"{col_key}"] = "up"
                if st.button("👎", key=f"down_{idx}"):
                    st.session_state[f"{col_key}"] = "down"

# ---------------- Risk Dashboard ----------------
with dashboard_tab:
    st.subheader("Graphical Risk Dashboard")
    if "report_df" in st.session_state and not st.session_state["report_df"].empty:
        dfx = st.session_state["report_df"]
        if "risk_level" in dfx.columns:
            rc = dfx["risk_level"].value_counts().reset_index()
            rc.columns = ["risk_level", "count"]
            fig = px.bar(rc, x="risk_level", y="count", color="risk_level",
                         color_discrete_map={"Critical": "red", "High": "orange", "Medium": "yellow", "Low": "green"},
                         title="Findings by Risk Level")
            st.plotly_chart(fig, use_container_width=True)
        if "status" in dfx.columns:
            sc = dfx["status"].value_counts().reset_index()
            sc.columns = ["status", "count"]
            fig2 = px.pie(sc, names="status", values="count", title="Open vs Closed Findings")
            st.plotly_chart(fig2, use_container_width=True)

# ---------------- Performance Metrics ----------------
with perf_tab:
    if "report_df" not in st.session_state:
        st.warning("Please run the Risk Assessment first to view performance metrics.")
    else:
        st.subheader("Performance Metrics")
        dfx = st.session_state["report_df"]

        # Total findings
        total_findings = len(dfx)
        st.markdown(f"**Total Findings:** {total_findings}")

        # High/Critical risks
        if "risk_level" in dfx.columns:
            high_critical_count = dfx[dfx["risk_level"].isin(["High", "Critical"])].shape[0]
            st.markdown(f"**High/Critical Risks:** {high_critical_count}")

            # Average risk score
            risk_map = {"Low": 1, "Medium": 2, "High": 3, "Critical": 4}
            dfx["risk_score"] = dfx["risk_level"].map(risk_map).fillna(0)
            avg_risk_score = dfx["risk_score"].mean()
            st.markdown(f"**Average Risk Score:** {avg_risk_score:.2f}")

        # Number of Q&A interactions
        num_qna = len([msg for msg in st.session_state.get("chat_history", []) if msg["role"] == "user"])
        st.markdown(f"**Number of Q&A Interactions:** {num_qna}")

        # Processing time for summary
        summary_time = st.session_state.get("summary_time_secs")
        if summary_time:
            st.markdown(f"**Processing Time for Summary:** {summary_time:.2f} secs")
        else:
            st.markdown("**Processing Time for Summary:** Not tracked")
